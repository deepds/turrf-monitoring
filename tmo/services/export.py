"""Выгрузка детализации цены в CSV и Excel.

Выгрузка содержит и включённые, и исключённые предложения с причиной
исключения: файл, в котором видно только вошедшее, не позволяет проверить
цифру — он позволяет лишь её повторить.

Секреты в выгрузку не попадают: в неё выводятся только поля, перечисленные
явно.
"""

from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any

from sqlalchemy.orm import Session

from tmo.services.showcase import metric_details, metric_offers

COLUMNS: list[tuple[str, str]] = [
    ("offer_id", "ID предложения"),
    ("is_included", "В расчёте"),
    ("exclusion_reason", "Причина исключения"),
    ("exclusion_detail", "Пояснение"),
    ("source_code", "Источник"),
    ("price", "Цена, ₽"),
    ("source_price", "Цена источника"),
    ("price_basis", "База цены"),
    ("currency", "Валюта"),
    ("route", "Маршрут / объект"),
    ("carrier", "Перевозчик"),
    ("vehicle", "Рейс / поезд"),
    ("car_type", "Тип вагона"),
    ("service_class", "Сервисный класс"),
    ("fare_family", "Тариф"),
    ("refundable", "Возвратный"),
    ("property_name", "Гостиница"),
    ("stars", "Звёзды"),
    ("property_type", "Тип объекта"),
    ("room_name", "Номер"),
    ("departure_at", "Отправление"),
    ("arrival_at", "Прибытие"),
    ("return_departure_at", "Обратное отправление"),
    ("check_in", "Заезд"),
    ("check_out", "Выезд"),
    ("nights", "Ночей"),
    ("fetched_at", "Получено"),
    ("validation_flags", "Отметки валидации"),
    ("deeplink", "Ссылка"),
]

PROVENANCE_COLUMNS: list[tuple[str, str]] = [
    ("source_attempt_id", "ID обращения"),
    ("raw_response_id", "ID сырого ответа"),
    ("raw_storage_ref", "Файл сырого ответа"),
    ("raw_sha256", "Хеш сырого ответа"),
    ("raw_page", "Страница выдачи"),
]


def _rows(session: Session, metric_id: int) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    details = metric_details(session, metric_id)
    offers = metric_offers(session, metric_id)
    return details, offers


def _flatten(offer: dict[str, Any]) -> dict[str, Any]:
    flat = {key: offer.get(key) for key, _ in COLUMNS}
    flat["validation_flags"] = ", ".join(offer.get("validation_flags") or [])
    flat["is_included"] = "да" if offer.get("is_included") else "нет"
    provenance = offer.get("provenance") or {}
    for key, _ in PROVENANCE_COLUMNS:
        flat[key] = provenance.get(key)
    return flat


def _header_lines(details: dict[str, Any]) -> list[list[Any]]:
    """Шапка выгрузки: по ней файл опознаётся без обращения к системе."""
    return [
        ["Метрика", details["metric_id"], "Тип", details["metric_type"]],
        [
            "Снимок",
            details["snapshot_date"],
            "Статус",
            details["snapshot_status"],
            "Расчёт",
            details["calculation_run_id"],
        ],
        ["Версия методики", details["methodology_version"], "Получено", details["fetched_at"]],
        [
            "Медиана",
            details["median_price"],
            "Минимум",
            details["min_price"],
            "Предложений",
            details["offers_count"],
            "Источников",
            details["sources_count"],
        ],
        [
            "Качество",
            details["quality_score"],
            "Уверенность",
            details["confidence_level"],
            "Частичная выборка",
            "да" if details["is_partial"] else "нет",
        ],
        ["Предупреждения", ", ".join(details["warning_codes"]) or "нет"],
        [],
    ]


def to_csv(session: Session, metric_id: int) -> tuple[str, bytes]:
    details, offers = _rows(session, metric_id)
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    for line in _header_lines(details):
        writer.writerow(line)
    writer.writerow([title for _, title in COLUMNS] + [title for _, title in PROVENANCE_COLUMNS])
    for offer in offers:
        flat = _flatten(offer)
        writer.writerow(
            [flat.get(key) for key, _ in COLUMNS] + [flat.get(key) for key, _ in PROVENANCE_COLUMNS]
        )
    filename = f"metric_{metric_id}_{details['snapshot_date']}.csv"
    # UTF-8 с BOM: без него Excel открывает кириллицу в неверной кодировке.
    return filename, b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def to_xlsx(session: Session, metric_id: int) -> tuple[str, bytes]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    details, offers = _rows(session, metric_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Детализация"

    bold = Font(bold=True)
    for line in _header_lines(details):
        sheet.append(line)
        if line:
            sheet.cell(row=sheet.max_row, column=1).font = bold

    header_row = sheet.max_row + 1
    titles = [title for _, title in COLUMNS] + [title for _, title in PROVENANCE_COLUMNS]
    sheet.append(titles)
    fill = PatternFill("solid", fgColor="EEF2F7")
    for index in range(1, len(titles) + 1):
        cell = sheet.cell(row=header_row, column=index)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    excluded_fill = PatternFill("solid", fgColor="FDECEC")
    for offer in offers:
        flat = _flatten(offer)
        sheet.append(
            [flat.get(key) for key, _ in COLUMNS] + [flat.get(key) for key, _ in PROVENANCE_COLUMNS]
        )
        if not offer.get("is_included"):
            # Исключённое видно глазом: файл должен объяснять, а не только
            # перечислять.
            for index in range(1, len(titles) + 1):
                sheet.cell(row=sheet.max_row, column=index).fill = excluded_fill

    for index, (_, title) in enumerate(COLUMNS + PROVENANCE_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(12, min(38, len(title) + 6))
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    stream = io.BytesIO()
    workbook.save(stream)
    filename = f"metric_{metric_id}_{details['snapshot_date']}.xlsx"
    return filename, stream.getvalue()


def timestamped(filename: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    name, _, extension = filename.rpartition(".")
    return f"{name}_{stamp}.{extension}"
